"""Export API tests — 202→complete, watermark from DB tier, spoof ignored, regen."""

from __future__ import annotations

import io
import time
import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from unittest.mock import patch

import pytest
from httpx import AsyncClient
from openpyxl import load_workbook

from app.db import SyncSessionLocal, set_rls_org_id
from app.main import app
from app.models.export import Export
from app.models.financial_statement import FinancialStatement
from app.models.organisation import Organisation
from app.models.statement_line_item import StatementLineItem
from app.models.trial_balance import TrialBalance
from app.routers.export import get_object_storage
from app.services.exporter import WATERMARK_TEXT, export_object_key
from tests.conftest import auth_headers


class InMemoryObjectStorage:
    """Test double for S3/R2 — captures bytes so watermark content can be asserted."""

    def __init__(self) -> None:
        self.objects: dict[str, bytes] = {}
        self.content_types: dict[str, str] = {}
        self.put_calls = 0

    def put_export(
        self,
        *,
        key: str,
        body: bytes,
        content_type: str,
        expires_at: datetime,
    ) -> None:
        self.objects[key] = body
        self.content_types[key] = content_type
        self.put_calls += 1

    def generate_signed_url(self, *, key: str, expires_in: int) -> str:
        return f"https://test-storage.local/{key}?sig=test&expires_in={expires_in}"

    def object_exists(self, *, key: str) -> bool:
        return key in self.objects


def _seed_tb_with_statements(
    *,
    org_id: uuid.UUID,
    client_id: uuid.UUID,
    period_end: date,
) -> uuid.UUID:
    tb_id = uuid.uuid4()
    with SyncSessionLocal() as session:
        set_rls_org_id(session, org_id)
        session.add(
            TrialBalance(
                id=tb_id,
                client_id=client_id,
                period_end=period_end,
                file_url=f"/tmp/{tb_id}.xlsx",
                file_type="xlsx",
                status="complete",
                currency="GBP",
                parsed_data={"rows": []},
            )
        )
        session.flush()
        for statement_type, lines in (
            (
                "SOPL",
                [
                    ("revenue", "Revenue", "10000.00", False, 0),
                    ("gross_profit", "Gross profit", "10000.00", True, 1),
                ],
            ),
            (
                "SOFP",
                [
                    ("cash", "Cash", "5000.00", False, 0),
                    ("total_assets", "Total assets", "5000.00", True, 1),
                ],
            ),
            (
                "SOCIE",
                [
                    (
                        "retained_earnings_closing",
                        "Retained earnings (closing)",
                        "5000.00",
                        True,
                        0,
                    ),
                ],
            ),
        ):
            fs = FinancialStatement(
                tb_id=tb_id,
                statement_type=statement_type,
                data={"lines": []},
            )
            session.add(fs)
            session.flush()
            for code, name, amount, is_subtotal, order in lines:
                session.add(
                    StatementLineItem(
                        statement_id=fs.id,
                        line_item_code=code,
                        line_item_name=name,
                        amount=Decimal(amount),
                        is_subtotal=is_subtotal,
                        display_order=order,
                    )
                )
        session.commit()
    return tb_id


def _set_org_tier(org_id: uuid.UUID, tier: str) -> None:
    with SyncSessionLocal() as session:
        set_rls_org_id(session, org_id)
        org = session.get(Organisation, org_id)
        assert org is not None
        org.subscription_tier = tier
        session.commit()


async def _poll_export_complete(
    api_client: AsyncClient,
    *,
    export_id: uuid.UUID,
    headers: dict[str, str],
) -> dict:
    body = None
    for _ in range(40):
        status_resp = await api_client.get(f"/exports/{export_id}", headers=headers)
        assert status_resp.status_code == 200, status_resp.text
        body = status_resp.json()
        if body["status"] in {"complete", "failed"}:
            return body
        time.sleep(0.05)
    pytest.fail(f"timed out waiting for export: {body}")


def _workbook_has_watermark(content: bytes) -> bool:
    workbook = load_workbook(io.BytesIO(content))
    for sheet_name in workbook.sheetnames:
        for row in workbook[sheet_name].iter_rows(max_row=12, max_col=2):
            for cell in row:
                if cell.value is not None and WATERMARK_TEXT in str(cell.value):
                    return True
    return False


@pytest.fixture
def memory_storage() -> InMemoryObjectStorage:
    storage = InMemoryObjectStorage()
    app.dependency_overrides[get_object_storage] = lambda: storage
    yield storage
    app.dependency_overrides.pop(get_object_storage, None)


@pytest.mark.asyncio
async def test_export_request_202_poll_complete(
    api_client: AsyncClient,
    provisioned_org: dict,
    memory_storage: InMemoryObjectStorage,
) -> None:
    headers = auth_headers(provisioned_org["token"])
    tb_id = _seed_tb_with_statements(
        org_id=provisioned_org["org_id"],
        client_id=provisioned_org["client_id"],
        period_end=date(2026, 7, 31),
    )

    accepted = await api_client.post(
        f"/trial-balances/{tb_id}/export",
        headers=headers,
        json={
            "format": "xlsx",
            "options": {
                "include_mapping_summary": True,
                "include_risk_report": True,
            },
        },
    )
    assert accepted.status_code == 202, accepted.text
    body = accepted.json()
    assert body["status"] == "pending"
    assert body["tb_id"] == str(tb_id)
    assert "export_id" in body
    assert "Processing will begin shortly" in body["message"]
    export_id = uuid.UUID(body["export_id"])

    final = await _poll_export_complete(
        api_client, export_id=export_id, headers=headers
    )
    assert final["status"] == "complete", final
    assert final["file_url"]
    assert memory_storage.put_calls >= 1

    download = await api_client.get(
        f"/exports/{export_id}/download",
        headers=headers,
        follow_redirects=False,
    )
    assert download.status_code == 302
    assert download.headers["location"].startswith("https://test-storage.local/")


@pytest.mark.asyncio
async def test_free_tier_export_is_watermarked_in_file_bytes(
    api_client: AsyncClient,
    provisioned_org: dict,
    memory_storage: InMemoryObjectStorage,
) -> None:
    """Watermark from DB subscription_tier=free — assert actual xlsx cell content."""
    org_id = provisioned_org["org_id"]
    _set_org_tier(org_id, "free")
    headers = auth_headers(provisioned_org["token"])
    tb_id = _seed_tb_with_statements(
        org_id=org_id,
        client_id=provisioned_org["client_id"],
        period_end=date(2026, 8, 31),
    )

    accepted = await api_client.post(
        f"/trial-balances/{tb_id}/export",
        headers=headers,
        json={"format": "xlsx"},
    )
    assert accepted.status_code == 202, accepted.text
    export_id = uuid.UUID(accepted.json()["export_id"])
    final = await _poll_export_complete(
        api_client, export_id=export_id, headers=headers
    )
    assert final["status"] == "complete"

    key = export_object_key(export_id, "xlsx")
    content = memory_storage.objects[key]
    assert _workbook_has_watermark(content) is True


@pytest.mark.asyncio
async def test_starter_tier_export_has_no_watermark(
    api_client: AsyncClient,
    provisioned_org: dict,
    memory_storage: InMemoryObjectStorage,
) -> None:
    org_id = provisioned_org["org_id"]
    _set_org_tier(org_id, "starter")
    headers = auth_headers(provisioned_org["token"])
    tb_id = _seed_tb_with_statements(
        org_id=org_id,
        client_id=provisioned_org["client_id"],
        period_end=date(2026, 9, 30),
    )

    accepted = await api_client.post(
        f"/trial-balances/{tb_id}/export",
        headers=headers,
        json={"format": "xlsx"},
    )
    assert accepted.status_code == 202, accepted.text
    export_id = uuid.UUID(accepted.json()["export_id"])
    final = await _poll_export_complete(
        api_client, export_id=export_id, headers=headers
    )
    assert final["status"] == "complete"

    key = export_object_key(export_id, "xlsx")
    content = memory_storage.objects[key]
    assert _workbook_has_watermark(content) is False


@pytest.mark.asyncio
async def test_request_body_tier_watermark_override_ignored(
    api_client: AsyncClient,
    provisioned_org: dict,
    memory_storage: InMemoryObjectStorage,
) -> None:
    """Client passes subscription_tier=scale + watermark=false; DB says free → still watermarked."""
    org_id = provisioned_org["org_id"]
    _set_org_tier(org_id, "free")
    headers = auth_headers(provisioned_org["token"])
    tb_id = _seed_tb_with_statements(
        org_id=org_id,
        client_id=provisioned_org["client_id"],
        period_end=date(2026, 10, 31),
    )

    accepted = await api_client.post(
        f"/trial-balances/{tb_id}/export",
        headers=headers,
        json={
            "format": "xlsx",
            "subscription_tier": "scale",
            "watermark": False,
            "options": {"include_mapping_summary": True, "include_risk_report": False},
        },
    )
    # Must not 422 (reject) and must not honour the spoofed tier.
    assert accepted.status_code == 202, accepted.text
    export_id = uuid.UUID(accepted.json()["export_id"])
    final = await _poll_export_complete(
        api_client, export_id=export_id, headers=headers
    )
    assert final["status"] == "complete"

    with SyncSessionLocal() as session:
        set_rls_org_id(session, org_id)
        org = session.get(Organisation, org_id)
        assert org is not None
        assert org.subscription_tier == "free"
        row = session.get(Export, export_id)
        assert row is not None
        # Spoof fields must not land in options JSONB.
        assert "subscription_tier" not in (row.options or {})
        assert "watermark" not in (row.options or {})

    key = export_object_key(export_id, "xlsx")
    assert _workbook_has_watermark(memory_storage.objects[key]) is True


@pytest.mark.asyncio
async def test_download_after_expiry_invokes_regenerate_export_if_missing(
    api_client: AsyncClient,
    provisioned_org: dict,
    memory_storage: InMemoryObjectStorage,
) -> None:
    headers = auth_headers(provisioned_org["token"])
    tb_id = _seed_tb_with_statements(
        org_id=provisioned_org["org_id"],
        client_id=provisioned_org["client_id"],
        period_end=date(2026, 11, 30),
    )

    accepted = await api_client.post(
        f"/trial-balances/{tb_id}/export",
        headers=headers,
        json={"format": "xlsx"},
    )
    assert accepted.status_code == 202, accepted.text
    export_id = uuid.UUID(accepted.json()["export_id"])
    final = await _poll_export_complete(
        api_client, export_id=export_id, headers=headers
    )
    assert final["status"] == "complete"
    puts_after_create = memory_storage.put_calls
    key = export_object_key(export_id, "xlsx")
    assert key in memory_storage.objects

    # Simulate 30-day lifecycle expiry — object gone from storage.
    del memory_storage.objects[key]
    assert memory_storage.object_exists(key=key) is False

    from app.services.exporter import regenerate_export_if_missing as real_regen

    with patch(
        "app.routers.export.regenerate_export_if_missing",
        side_effect=real_regen,
    ) as regen_spy:
        download = await api_client.get(
            f"/exports/{export_id}/download",
            headers=headers,
            follow_redirects=False,
        )

    assert download.status_code == 302, download.text
    assert regen_spy.called
    assert memory_storage.put_calls > puts_after_create
    assert key in memory_storage.objects


@pytest.mark.asyncio
async def test_download_while_pending_returns_conflict_not_broken_url(
    api_client: AsyncClient,
    provisioned_org: dict,
    memory_storage: InMemoryObjectStorage,
) -> None:
    headers = auth_headers(provisioned_org["token"])
    tb_id = _seed_tb_with_statements(
        org_id=provisioned_org["org_id"],
        client_id=provisioned_org["client_id"],
        period_end=date(2026, 12, 31),
    )
    # Insert a stuck pending export without running the job.
    export_id = uuid.uuid4()
    with SyncSessionLocal() as session:
        set_rls_org_id(session, provisioned_org["org_id"])
        session.add(
            Export(
                id=export_id,
                tb_id=tb_id,
                format="xlsx",
                status="pending",
                options={},
            )
        )
        session.commit()

    download = await api_client.get(
        f"/exports/{export_id}/download",
        headers=headers,
        follow_redirects=False,
    )
    assert download.status_code == 409
    assert "not ready" in download.json()["detail"].lower()
