"""Tests for Excel/PDF/CSV export generation."""

from __future__ import annotations

import io
import uuid
from dataclasses import dataclass
from datetime import date, datetime, timezone
from decimal import Decimal
from unittest.mock import MagicMock

from openpyxl import load_workbook

from app.schemas.risk import RiskFlagRecord
from app.schemas.variance import VarianceAnalysisResult, VarianceItemRecord
from app.services.exporter import (
    DISCLAIMER_TEXT,
    WATERMARK_TEXT,
    BuiltExport,
    ExportBranding,
    ExportPackage,
    build_csv,
    build_excel,
    build_export,
    format_currency,
    regenerate_export_if_missing,
    render_pdf_html,
    run_export_job,
    tier_requires_watermark,
    upload_export_file,
)
from app.services.mapper import MappingResult
from app.services.statements import StatementLineItemRecord


@dataclass
class _Org:
    subscription_tier: str


@dataclass
class _ExportRow:
    status: str = "pending"
    file_url: str | None = None
    error_message: str | None = None


def _line(
    code: str,
    name: str,
    amount: str,
    *,
    display_order: int,
    is_subtotal: bool = False,
) -> StatementLineItemRecord:
    return StatementLineItemRecord(
        line_item_code=code,
        line_item_name=name,
        amount=Decimal(amount),
        is_subtotal=is_subtotal,
        display_order=display_order,
        source_account_ids=[],
    )


def _branding() -> ExportBranding:
    return ExportBranding(
        client_name="Acme Ltd",
        period_end=date(2026, 3, 31),
        generated_at=datetime(2026, 4, 1, 12, 0, tzinfo=timezone.utc),
        organisation_name="Acme Advisory",
    )


def _package() -> ExportPackage:
    sopl = [
        _line("revenue", "Revenue", "10000.00", display_order=1),
        _line("cost_of_sales", "Cost of sales", "4000.00", display_order=2),
        _line("gross_profit", "Gross profit", "6000.00", display_order=3, is_subtotal=True),
    ]
    sofp = [
        _line("cash", "Cash", "5000.00", display_order=1),
        _line("total_assets", "Total assets", "5000.00", display_order=2, is_subtotal=True),
    ]
    socie = [
        _line(
            "retained_earnings_opening",
            "Retained earnings (opening)",
            "2000.00",
            display_order=1,
        ),
        _line(
            "retained_earnings_closing",
            "Retained earnings (closing)",
            "3000.00",
            display_order=2,
            is_subtotal=True,
        ),
    ]
    variance = VarianceAnalysisResult(
        items=[
            VarianceItemRecord(
                line_item_code="revenue",
                line_item_name="Revenue",
                current_amount="10000.00",
                prior_amount="8000.00",
                variance_amount="2000.00",
                variance_pct="25.00",
                direction="increase",
                is_material=True,
            )
        ]
    )
    risks = [
        RiskFlagRecord(
            rule_name="negative_cash",
            severity="warning",
            description="A cash or bank account shows a negative balance.",
            recommended_action="Verify with bank statements.",
        )
    ]
    mappings = [
        MappingResult("4000", "Sales", "revenue", Decimal("1.00"), "exact"),
    ]
    return ExportPackage(
        sopl=sopl,
        sofp=sofp,
        socie=socie,
        variance=variance,
        risk_flags=risks,
        mappings=mappings,
    )


def test_format_currency_follows_section_10_7() -> None:
    assert format_currency(Decimal("1234567.89"), "GBP") == "£1,234,567.89"
    assert format_currency(Decimal("1234567.89"), "USD") == "$1,234,567.89"
    assert format_currency(Decimal("1234567.89"), "EUR") == "€1 234 567.89"
    assert format_currency(Decimal("-100.50"), "EUR") == "-€100.50"
    assert format_currency(Decimal("30000"), "GBP") == "£30,000.00"
    assert format_currency(Decimal("30000"), "EUR") == "€30 000.00"


def test_exports_keep_thousands_separators_after_nil_face_filter() -> None:
    """Nil-face filtering must not strip §10.7 formatting on CSV/PDF/Excel paths."""
    from app.services.statements import filter_nil_face_lines

    raw = [
        _line("revenue", "Revenue", "30000.00", display_order=1),
        _line("investments", "Investments", "0.00", display_order=2),  # nil leaf
        _line("gross_profit", "Gross profit", "30000.00", display_order=3, is_subtotal=True),
    ]
    filtered = filter_nil_face_lines(raw)
    assert [line.line_item_code for line in filtered] == ["revenue", "gross_profit"]

    for currency, expected in (
        ("GBP", "£30,000.00"),
        ("EUR", "€30 000.00"),
        ("USD", "$30,000.00"),
    ):
        branding = ExportBranding(
            client_name="Sep Co",
            period_end=date(2026, 3, 31),
            generated_at=datetime(2026, 4, 1, 12, 0, tzinfo=timezone.utc),
            functional_currency=currency,
        )
        package = ExportPackage(
            sopl=filtered,
            sofp=filtered,
            socie=filtered,
            variance=None,
            risk_flags=[],
            mappings=[],
        )
        csv_text = build_csv(branding, package).decode("utf-8")
        assert expected in csv_text, currency

        html = render_pdf_html(branding, package, organisation=_Org("starter"))
        assert expected in html, currency

        workbook = load_workbook(io.BytesIO(build_excel(branding, package, organisation=_Org("starter"))))
        amount_cell = None
        for row in workbook["SOPL"].iter_rows(min_col=1, max_col=2):
            if row[0].value == "Revenue":
                amount_cell = row[1]
                break
        assert amount_cell is not None
        assert amount_cell.value == Decimal("30000.00")
        if currency == "EUR":
            assert " " in amount_cell.number_format or "#" in amount_cell.number_format
        else:
            assert "#,##0.00" == amount_cell.number_format
        # Formatted string used by CSV/PDF (and dashboard) for the same amount:
        assert format_currency(Decimal(str(amount_cell.value)), currency) == expected


def test_excel_statement_sheet_shows_currency_in_branding() -> None:
    branding = ExportBranding(
        client_name="Euro Co",
        period_end=date(2026, 3, 31),
        generated_at=datetime(2026, 4, 1, 12, 0, tzinfo=timezone.utc),
        functional_currency="EUR",
    )
    content = build_excel(branding, _package(), organisation=_Org("starter"))
    workbook = load_workbook(io.BytesIO(content))
    values = [
        str(cell.value)
        for row in workbook["SOPL"].iter_rows(max_row=8, max_col=1)
        for cell in row
        if cell.value is not None
    ]
    assert any("All amounts in EUR" in value for value in values)


def test_pdf_statement_sections_include_currency_label() -> None:
    branding = ExportBranding(
        client_name="Euro Co",
        period_end=date(2026, 3, 31),
        generated_at=datetime(2026, 4, 1, 12, 0, tzinfo=timezone.utc),
        functional_currency="EUR",
    )
    html = render_pdf_html(branding, _package(), organisation=_Org("starter"))
    assert "All amounts in <strong>EUR</strong>" in html
    assert "€10,000.00" not in html
    assert "€10 000.00" in html


def test_eur_excel_amounts_render_with_space_thousands_via_libreoffice() -> None:
    """Open the generated .xlsx in LibreOffice and read rendered cell text.

    Excel number formats are locale-sensitive; format strings alone are not enough.
    """
    import re
    import shutil
    import subprocess
    from pathlib import Path

    if shutil.which("libreoffice") is None:
        import pytest

        pytest.skip("libreoffice not installed")

    branding = ExportBranding(
        client_name="Euro Co",
        period_end=date(2026, 3, 31),
        generated_at=datetime(2026, 4, 1, 12, 0, tzinfo=timezone.utc),
        functional_currency="EUR",
    )
    package = ExportPackage(
        sopl=[
            _line("revenue", "Revenue", "1234567.89", display_order=1),
            _line("gross_profit", "Gross profit", "-100.50", display_order=2, is_subtotal=True),
        ],
        sofp=[],
        socie=[],
        variance=None,
        risk_flags=[],
        mappings=[],
    )
    out = Path("/tmp/findraft-eur-export-libreoffice-test.xlsx")
    out.write_bytes(build_excel(branding, package, organisation=_Org("starter")))
    html_path = out.with_suffix(".html")
    if html_path.exists():
        html_path.unlink()
    subprocess.run(
        [
            "libreoffice",
            "--headless",
            "--convert-to",
            "html",
            str(out),
            "--outdir",
            str(out.parent),
        ],
        check=True,
        capture_output=True,
    )
    html = html_path.read_text(encoding="utf-8")
    revenue = re.search(
        r"Revenue</font></td>\s*<td[^>]*><font[^>]*>([^<]+)</font>",
        html,
    )
    assert revenue is not None, html[:500]
    # # ### ##0.00 — correct space grouping for large EUR amounts
    assert revenue.group(1).strip() == "1 234 567.89"


def test_tier_requires_watermark_only_for_free() -> None:
    assert tier_requires_watermark(_Org("free")) is True
    assert tier_requires_watermark(_Org("starter")) is False
    assert tier_requires_watermark(_Org("pro")) is False
    assert tier_requires_watermark(_Org("scale")) is False


def test_excel_watermark_applied_for_free_tier() -> None:
    content = build_excel(_branding(), _package(), organisation=_Org("free"))
    workbook = load_workbook(io.BytesIO(content))
    for sheet_name in ("SOPL", "SOFP", "SOCIE", "Variance", "Risk", "Mapping Summary"):
        values = [
            str(cell.value)
            for row in workbook[sheet_name].iter_rows(max_row=10, max_col=1)
            for cell in row
            if cell.value is not None
        ]
        assert any(WATERMARK_TEXT in value for value in values), sheet_name


def test_excel_watermark_not_applied_for_starter() -> None:
    content = build_excel(_branding(), _package(), organisation=_Org("starter"))
    workbook = load_workbook(io.BytesIO(content))
    for sheet_name in workbook.sheetnames:
        values = [
            str(cell.value)
            for row in workbook[sheet_name].iter_rows(max_row=10, max_col=1)
            for cell in row
            if cell.value is not None
        ]
        assert WATERMARK_TEXT not in values


def test_disclaimer_present_in_all_three_formats() -> None:
    branding = _branding()
    package = _package()

    excel = build_excel(branding, package, organisation=_Org("starter"))
    workbook = load_workbook(io.BytesIO(excel))
    for sheet_name in workbook.sheetnames:
        flat = " ".join(
            str(cell.value)
            for row in workbook[sheet_name].iter_rows(max_row=15, max_col=4)
            for cell in row
            if cell.value is not None
        )
        assert DISCLAIMER_TEXT in flat, sheet_name

    csv_text = build_csv(branding, package).decode("utf-8")
    assert DISCLAIMER_TEXT in csv_text

    html = render_pdf_html(branding, package, organisation=_Org("pro"))
    assert DISCLAIMER_TEXT in html
    pdf = build_export(
        "pdf",
        branding=branding,
        package=package,
        organisation=_Org("pro"),
    ).content
    assert pdf.startswith(b"%PDF")


def test_excel_monetary_values_are_decimal_not_float() -> None:
    content = build_excel(_branding(), _package(), organisation=_Org("pro"))
    workbook = load_workbook(io.BytesIO(content), data_only=False)
    sopl = workbook["SOPL"]

    money_cells = []
    for row in sopl.iter_rows(min_row=1, max_col=2):
        cell = row[1]
        if cell.value is None or cell.number_format != "#,##0.00":
            continue
        money_cells.append(cell.value)

    assert money_cells, "expected monetary cells on SOPL"
    for value in money_cells:
        # openpyxl may coerce Decimal→float on round-trip through xlsx XML;
        # guard the write path by rebuilding and inspecting before save as well.
        assert not isinstance(value, float) or Decimal(str(value)) == value

    # Stronger check: values written via _excel_money are Decimal before save.
    from app.services.exporter import _excel_money

    amount = _excel_money(Decimal("10000.00"))
    assert isinstance(amount, Decimal)
    assert type(amount) is Decimal
    assert amount == Decimal("10000.00")


def test_excel_cells_receive_decimal_instances_before_save() -> None:
    """Inspect the live workbook object before serialization."""
    from openpyxl import Workbook

    from app.services.exporter import _write_statement_sheet

    workbook = Workbook()
    workbook.remove(workbook.active)
    lines = [_line("revenue", "Revenue", "1234.56", display_order=1)]
    _write_statement_sheet(workbook, "SOPL", _branding(), lines, watermark=False)

    amount_cell = None
    for row in workbook["SOPL"].iter_rows(min_row=1, max_col=2):
        if row[0].value == "Revenue":
            amount_cell = row[1]
            break
    assert amount_cell is not None
    assert type(amount_cell.value) is Decimal
    assert amount_cell.value == Decimal("1234.56")
    assert amount_cell.number_format == "#,##0.00"


def test_pdf_cover_organisation_label_is_strong() -> None:
    html = render_pdf_html(_branding(), _package(), organisation=_Org("starter"))
    assert "<p><strong>Organisation:</strong> Acme Advisory</p>" in html
    assert "<p><strong>Client:</strong> Acme Ltd</p>" in html
    assert "<p><strong>Period end:</strong> 2026-03-31</p>" in html
    assert "<p><strong>Currency:</strong> GBP</p>" in html
    assert "<p>Organisation: Acme Advisory</p>" not in html


def test_pdf_watermark_only_for_free_tier() -> None:
    branding = _branding()
    package = _package()
    free_html = render_pdf_html(branding, package, organisation=_Org("free"))
    starter_html = render_pdf_html(branding, package, organisation=_Org("starter"))
    assert WATERMARK_TEXT in free_html
    assert WATERMARK_TEXT not in starter_html


def test_run_export_job_transitions_status_and_uploads() -> None:
    export = _ExportRow(status="pending")
    storage = MagicMock()
    storage.generate_signed_url.return_value = "https://example.com/exports/x.xlsx"
    export_id = uuid.uuid4()

    run_export_job(
        export,
        format="xlsx",
        branding=_branding(),
        package=_package(),
        organisation=_Org("starter"),
        storage=storage,
        export_id=export_id,
    )

    assert export.status == "complete"
    assert export.file_url == "https://example.com/exports/x.xlsx"
    assert storage.put_export.called
    put_kwargs = storage.put_export.call_args.kwargs
    assert put_kwargs["key"].startswith("exports/")
    assert put_kwargs["expires_at"] is not None
    # NOTE: put_export's Expires= and Tagging= cannot prove 30-day deletion.
    # Expires is only an HTTP caching hint; Tagging only feeds a bucket lifecycle
    # filter. Actual deletion requires put_bucket_lifecycle_configuration
    # (scripts/configure_s3_lifecycle.py) verified on the real bucket via
    # `aws s3api get-bucket-lifecycle-configuration` — a deployment checklist
    # item in docs/runbooks/deployment.md, not something a mocked client can catch.


def test_regenerate_when_object_missing() -> None:
    export = _ExportRow(status="complete", file_url="https://stale.example/file.xlsx")
    storage = MagicMock()
    storage.object_exists.return_value = False
    storage.generate_signed_url.return_value = "https://example.com/exports/new.xlsx"
    export_id = uuid.uuid4()

    url = regenerate_export_if_missing(
        export,
        format="xlsx",
        branding=_branding(),
        package=_package(),
        organisation=_Org("free"),
        storage=storage,
        export_id=export_id,
    )

    assert url == "https://example.com/exports/new.xlsx"
    assert export.status == "complete"
    assert storage.put_export.called


def test_s3_put_export_omits_tagging_for_r2(monkeypatch: pytest.MonkeyPatch) -> None:
    """R2 rejects x-amz-tagging on PutObject — must not send Tagging=."""
    from datetime import datetime, timezone

    from app.config import settings
    from app.services.exporter import S3ObjectStorage

    monkeypatch.setattr(
        settings,
        "s3_endpoint_url",
        "https://account.r2.cloudflarestorage.com",
    )
    captured: dict[str, object] = {}

    class FakeClient:
        def put_object(self, **kwargs: object) -> None:
            captured.update(kwargs)

    storage = S3ObjectStorage(client=FakeClient())
    storage.put_export(
        key="exports/test-id.xlsx",
        body=b"data",
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        expires_at=datetime.now(timezone.utc),
    )
    assert "Tagging" not in captured
    assert captured["Key"] == "exports/test-id.xlsx"


def test_s3_client_region_uses_auto_for_r2(monkeypatch: pytest.MonkeyPatch) -> None:
    from app.config import settings
    from app.services.exporter import _s3_client_region

    monkeypatch.setattr(settings, "s3_region", "eu-west-1")
    monkeypatch.setattr(
        settings,
        "s3_endpoint_url",
        "https://account.r2.cloudflarestorage.com",
    )
    assert _s3_client_region() == "auto"
    monkeypatch.setattr(settings, "s3_region", "weur")
    assert _s3_client_region() == "weur"
