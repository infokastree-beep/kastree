"""Tests for trial balance parser."""

from __future__ import annotations

import logging
from decimal import Decimal
from io import BytesIO

import openpyxl
import pytest

from app.services.parser import (
    AmbiguousCurrencyError,
    OrphanedAmountError,
    ParseError,
    TBRow,
    UnbalancedTrialBalanceError,
    parse_monetary,
    parse_tb_file,
)


def _xlsx_bytes(build_sheet) -> bytes:
    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    build_sheet(worksheet)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _multi_sheet_xlsx_bytes(*sheet_builders) -> bytes:
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for index, builder in enumerate(sheet_builders):
        worksheet = workbook.create_sheet(title=f"Sheet{index + 1}")
        builder(worksheet)
    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def test_parse_standard_four_column_format() -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["1000", "Cash", "10000.00", "0.00"])
        ws.append(["2000", "Sales", "0.00", "10000.00"])

    rows = parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    assert rows == [
        TBRow(
            account_code="1000",
            account_name="Cash",
            debit=Decimal("10000.00"),
            credit=Decimal("0.00"),
            net_balance=Decimal("10000.00"),
            currency="GBP",
            row_index=2,
        ),
        TBRow(
            account_code="2000",
            account_name="Sales",
            debit=Decimal("0.00"),
            credit=Decimal("10000.00"),
            net_balance=Decimal("-10000.00"),
            currency="GBP",
            row_index=3,
        ),
    ]


def test_parse_single_balance_column_format() -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Balance"])
        ws.append(["1000", "Cash", "10000.00"])
        ws.append(["2000", "Sales", "-10000.00"])

    rows = parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="EUR")

    assert rows[0].debit == Decimal("10000.00")
    assert rows[0].credit == Decimal("0.00")
    assert rows[1].debit == Decimal("0.00")
    assert rows[1].credit == Decimal("10000.00")
    assert rows[1].net_balance == Decimal("-10000.00")


def test_parse_multi_tab_workbook_falls_back_to_second_sheet() -> None:
    def build_sparse(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["9999", "Placeholder", "1.00", "1.00"])

    def build_full(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["1000", "Cash", "500.00", "0.00"])
        ws.append(["1100", "Debtors", "250.00", "0.00"])
        ws.append(["2000", "Sales", "0.00", "750.00"])

    rows = parse_tb_file(
        _multi_sheet_xlsx_bytes(build_sparse, build_full),
        filename="tb.xlsx",
        functional_currency="GBP",
    )

    assert len(rows) == 3
    assert [row.account_code for row in rows] == ["1000", "1100", "2000"]


def test_parse_merged_cells_forward_fills_and_logs_warning(caplog: pytest.LogCaptureFixture) -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["1000", "Cash", "5000.00", "0.00"])
        ws.append([None, "Petty Cash", "5000.00", "0.00"])
        ws.append(["2000", "Sales", "0.00", "10000.00"])
        ws.merge_cells("A2:A3")

    with caplog.at_level(logging.WARNING):
        rows = parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    assert any("Merged cells detected" in record.message for record in caplog.records)
    assert rows[0].account_code == "1000"
    assert rows[1].account_code == "1000"
    assert rows[1].account_name == "Petty Cash"


def test_parse_skips_header_totals_and_blank_rows() -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["Account", "Description", "Debit", "Credit"])
        ws.append(["", "", "", ""])
        ws.append(["1000", "Cash", "6000.00", "0.00"])
        ws.append(["", "Total Assets", "6000.00", "0.00"])
        ws.append(["2000", "Sales", "0.00", "6000.00"])

    rows = parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    assert len(rows) == 2
    assert rows[0].account_name == "Cash"
    assert rows[1].account_name == "Sales"


def test_parse_strips_currency_symbols() -> None:
    assert parse_monetary("€1,234.56") == Decimal("1234.56")
    assert parse_monetary("$0.00") == Decimal("0")

    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["1000", "Cash", "£1,000.00", "£0.00"])
        ws.append(["2000", "Sales", "£0.00", "£1,000.00"])

    rows = parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    assert rows[0].debit == Decimal("1000.00")
    assert rows[1].credit == Decimal("1000.00")
    assert rows[0].currency == "GBP"


def test_parse_parentheses_negative_amounts() -> None:
    assert parse_monetary("(1,000.00)") == Decimal("-1000.00")

    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Balance"])
        ws.append(["1000", "Cash", "1000.00"])
        ws.append(["2000", "Accrual", "(1000.00)"])

    rows = parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    assert rows[1].net_balance == Decimal("-1000.00")
    assert rows[1].credit == Decimal("1000.00")


def test_parse_rejects_unbalanced_trial_balance_with_difference() -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["1000", "Cash", "10000.00", "0.00"])
        ws.append(["2000", "Sales", "0.00", "9999.00"])

    with pytest.raises(UnbalancedTrialBalanceError) as exc_info:
        parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    error = exc_info.value
    assert error.total_debits == Decimal("10000.00")
    assert error.total_credits == Decimal("9999.00")
    assert error.difference == Decimal("1.00")
    assert "Difference: 1.00" in str(error)


def test_parse_raises_on_ambiguous_currency_symbols(
    caplog: pytest.LogCaptureFixture,
) -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["1000", "Cash £500", "500.00", "0.00"])
        ws.append(["2000", "Sales €500", "0.00", "500.00"])

    with caplog.at_level(logging.WARNING):
        with pytest.raises(AmbiguousCurrencyError) as exc_info:
            parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    assert exc_info.value.symbols == frozenset({"£", "€"})
    assert any("Ambiguous currency symbols detected" in record.message for record in caplog.records)


def test_parse_rejects_non_numeric_monetary_cells() -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["1000", "Cash", "not-a-number", "0.00"])
        ws.append(["2000", "Sales", "0.00", "0.00"])

    with pytest.raises(ParseError) as exc_info:
        parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    message = str(exc_info.value)
    assert "Non-numeric monetary value" in message
    assert "row 2" in message
    assert "debit" in message
    assert "not-a-number" in message


def test_parse_raises_orphaned_amount_error_for_unidentified_monetary_row() -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Debit", "Credit"])
        ws.append(["", "", "500.00", "0.00"])
        ws.append(["1000", "Cash", "500.00", "0.00"])

    with pytest.raises(OrphanedAmountError) as exc_info:
        parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    error = exc_info.value
    assert error.row_index == 2
    assert error.amount == Decimal("500.00")
    assert error.column_name == "debit"
    assert "row 2" in str(error)
    assert "500.00" in str(error)


def test_parse_csv_four_column_format() -> None:
    csv_content = (
        "Account Code,Account Name,Debit,Credit\n"
        "1000,Cash,100.00,0.00\n"
        "2000,Sales,0.00,100.00\n"
    ).encode()

    rows = parse_tb_file(csv_content, filename="tb.csv", functional_currency="GBP")

    assert len(rows) == 2
    assert rows[0].account_code == "1000"


def test_parse_uses_explicit_currency_column() -> None:
    def build(ws) -> None:
        ws.append(["Account Code", "Account Name", "Currency", "Debit", "Credit"])
        ws.append(["1000", "Cash", "USD", "100.00", "0.00"])
        ws.append(["2000", "Sales", "USD", "0.00", "100.00"])

    rows = parse_tb_file(_xlsx_bytes(build), filename="tb.xlsx", functional_currency="GBP")

    assert rows[0].currency == "USD"
    assert rows[1].currency == "USD"
