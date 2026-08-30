"""Trial balance file parser — pandas-based, Decimal-only arithmetic."""

from __future__ import annotations

import logging
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation
from io import BytesIO
from typing import BinaryIO, Literal

import openpyxl
import pandas as pd
from zipfile import BadZipFile

logger = logging.getLogger(__name__)

TOLERANCE = Decimal("0.01")
MIN_DATA_ROWS = 3
SCAN_CELL_LIMIT = 100

HEADER_FIRST_CELL_KEYWORDS = (
    "account",
    "code",
    "description",
    "debit",
    "credit",
    "balance",
    "currency",
    "name",
)
TOTALS_NAME_KEYWORDS = ("total", "balance", "sum")

SYMBOL_TO_CURRENCY: dict[str, str] = {
    "£": "GBP",
    "€": "EUR",
    "$": "USD",
}
ISO_CURRENCIES = frozenset({"GBP", "EUR", "USD"})

TBFormat = Literal["four_column", "single_balance"]


@dataclass(frozen=True)
class TBRow:
    account_code: str
    account_name: str
    debit: Decimal
    credit: Decimal
    net_balance: Decimal
    currency: str
    row_index: int


class ParseError(Exception):
    """Raised when a cell cannot be parsed as a monetary value."""


class OrphanedAmountError(Exception):
    """Raised when a row has monetary values but no account identifiers."""

    def __init__(self, row_index: int, amount: Decimal, column_name: str) -> None:
        self.row_index = row_index
        self.amount = amount
        self.column_name = column_name
        super().__init__(
            f"Monetary value {amount} found at row {row_index} in column "
            f"{column_name!r} without an account code or account name."
        )


class AmbiguousCurrencyError(Exception):
    """Raised when multiple currency symbols are detected in the file."""

    def __init__(self, symbols: frozenset[str]) -> None:
        self.symbols = symbols
        symbol_list = ", ".join(sorted(symbols))
        super().__init__(
            f"Multiple currency symbols detected ({symbol_list}). "
            "Confirm currency in the upload UI before parsing proceeds."
        )


class UnbalancedTrialBalanceError(Exception):
    """Raised when total debits and credits differ beyond tolerance."""

    def __init__(self, total_debits: Decimal, total_credits: Decimal) -> None:
        self.total_debits = total_debits
        self.total_credits = total_credits
        self.difference = abs(total_debits - total_credits)
        super().__init__(
            f"Trial balance is unbalanced: total debits ({total_debits}) do not equal "
            f"total credits ({total_credits}). Difference: {self.difference}"
        )


def decimal_eq(a: Decimal, b: Decimal) -> bool:
    return abs(a - b) <= TOLERANCE


def parse_monetary(
    value: object,
    *,
    row_index: int | None = None,
    column_name: str | None = None,
) -> Decimal:
    """Parse a monetary cell value into Decimal, never float."""
    if value is None:
        return Decimal("0")

    if isinstance(value, float) and pd.isna(value):
        return Decimal("0")

    text = str(value).strip()
    if text == "" or text.lower() in {"nan", "none"}:
        return Decimal("0")

    negative = False
    if text.startswith("(") and text.endswith(")"):
        negative = True
        text = text[1:-1].strip()

    for symbol in SYMBOL_TO_CURRENCY:
        text = text.replace(symbol, "")
    text = text.replace(",", "").strip()

    for code in ISO_CURRENCIES:
        if text.upper().endswith(code):
            text = text[: -len(code)].strip()

    if text in {"", "-"}:
        return Decimal("0")

    try:
        amount = Decimal(text)
    except InvalidOperation as exc:
        location = _format_cell_location(row_index=row_index, column_name=column_name)
        raise ParseError(f"Non-numeric monetary value{location}: {value!r}") from exc

    return -amount if negative else amount


def _format_cell_location(
    *,
    row_index: int | None,
    column_name: str | None,
) -> str:
    if row_index is not None and column_name is not None:
        return f" at row {row_index}, column {column_name!r}"
    if row_index is not None:
        return f" at row {row_index}"
    if column_name is not None:
        return f" in column {column_name!r}"
    return ""


def parse_tb_file(
    file: BinaryIO | bytes,
    *,
    filename: str,
    functional_currency: str = "GBP",
) -> list[TBRow]:
    """Parse an uploaded .xlsx or .csv trial balance file into TBRow objects."""
    content = file if isinstance(file, bytes) else file.read()
    extension = filename.rsplit(".", 1)[-1].lower()

    if extension == "csv":
        dataframe = _read_csv(content)
        rows = _parse_dataframe(dataframe, functional_currency=functional_currency)
    elif extension == "xlsx":
        prepared = _prepare_xlsx(content)
        rows = _parse_xlsx(prepared, functional_currency=functional_currency)
    else:
        raise ParseError(f"Unsupported file type: {extension!r}. Expected .xlsx or .csv.")

    _validate_balanced(rows)
    return rows


def _read_csv(content: bytes) -> pd.DataFrame:
    return pd.read_csv(BytesIO(content), header=None, dtype=str, keep_default_na=False)


def _prepare_xlsx(content: bytes) -> bytes:
    try:
        workbook = openpyxl.load_workbook(BytesIO(content), data_only=True)
    except BadZipFile as exc:
        raise ParseError(
            "This file is not a valid Excel (.xlsx) workbook. "
            "Export your trial balance as .xlsx or .csv from your accounting software."
        ) from exc
    merged_row_ranges: list[str] = []

    for worksheet in workbook.worksheets:
        for merged_range in list(worksheet.merged_cells.ranges):
            min_row = merged_range.min_row
            max_row = merged_range.max_row
            min_col = merged_range.min_col
            max_col = merged_range.max_col
            top_left_value = worksheet.cell(min_row, min_col).value
            worksheet.unmerge_cells(str(merged_range))
            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    worksheet.cell(row, col).value = top_left_value
            merged_row_ranges.append(f"{min_row}-{max_row}")

    if merged_row_ranges:
        logger.warning(
            "Merged cells detected in rows %s. Values inferred.",
            ", ".join(merged_row_ranges),
        )

    buffer = BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


def _parse_xlsx(content: bytes, *, functional_currency: str) -> list[TBRow]:
    excel = pd.ExcelFile(BytesIO(content), engine="openpyxl")
    sheet_names = excel.sheet_names
    if not sheet_names:
        raise ParseError("Workbook contains no worksheets.")

    first_sheet = sheet_names[0]
    first_df = pd.read_excel(
        excel, sheet_name=first_sheet, header=None, dtype=str, keep_default_na=False
    )
    first_rows = _parse_dataframe(first_df, functional_currency=functional_currency)
    if len(first_rows) >= MIN_DATA_ROWS or len(sheet_names) == 1:
        return first_rows

    second_sheet = sheet_names[1]
    logger.info(
        "First worksheet %r has fewer than %d data rows; using second worksheet %r.",
        first_sheet,
        MIN_DATA_ROWS,
        second_sheet,
    )
    second_df = pd.read_excel(
        excel, sheet_name=second_sheet, header=None, dtype=str, keep_default_na=False
    )
    return _parse_dataframe(second_df, functional_currency=functional_currency)


def _parse_dataframe(
    dataframe: pd.DataFrame,
    *,
    functional_currency: str,
) -> list[TBRow]:
    if dataframe.empty:
        return []

    working = dataframe.fillna("").astype(str)
    header_row_index = _find_header_row(working)
    if header_row_index is not None:
        headers = [_normalize_header(value) for value in working.iloc[header_row_index].tolist()]
        body = working.iloc[header_row_index + 1 :].reset_index(drop=True)
        body.columns = _make_unique_columns(headers, width=len(body.columns))
    else:
        body = working.reset_index(drop=True)
        body.columns = [f"col_{index}" for index in range(len(body.columns))]

    column_map, tb_format = _detect_columns(list(body.columns))
    default_currency, per_row_currency = _detect_currency(
        body,
        column_map=column_map,
        functional_currency=functional_currency,
    )

    rows: list[TBRow] = []
    base_row = (header_row_index + 1) if header_row_index is not None else 0
    for offset, series in body.iterrows():
        spreadsheet_row_index = int(offset) + base_row + 1

        account_code = _cell_text(series, column_map["account_code"])
        account_name = _cell_text(series, column_map["account_name"])

        if _is_header_row(account_code):
            continue

        if not account_code and not account_name:
            orphaned = _find_orphaned_amount(
                series,
                column_map,
                tb_format,
                row_index=spreadsheet_row_index,
            )
            if orphaned is None:
                continue
            column_name, amount = orphaned
            raise OrphanedAmountError(spreadsheet_row_index, amount, column_name)

        if _is_totals_row(account_name):
            continue

        if tb_format == "four_column":
            debit = parse_monetary(
                series[column_map["debit"]],
                row_index=spreadsheet_row_index,
                column_name=column_map["debit"],
            )
            credit = parse_monetary(
                series[column_map["credit"]],
                row_index=spreadsheet_row_index,
                column_name=column_map["credit"],
            )
        else:
            balance = parse_monetary(
                series[column_map["balance"]],
                row_index=spreadsheet_row_index,
                column_name=column_map["balance"],
            )
            if balance >= Decimal("0"):
                debit, credit = balance, Decimal("0")
            else:
                debit, credit = Decimal("0"), abs(balance)

        net_balance = debit - credit
        currency = (
            per_row_currency(spreadsheet_row_index, series)
            if per_row_currency is not None
            else default_currency
        )

        rows.append(
            TBRow(
                account_code=account_code,
                account_name=account_name,
                debit=debit,
                credit=credit,
                net_balance=net_balance,
                currency=currency,
                row_index=spreadsheet_row_index,
            )
        )

    return rows


def _find_header_row(dataframe: pd.DataFrame) -> int | None:
    for index in range(min(10, len(dataframe))):
        row_values = [_normalize_header(value) for value in dataframe.iloc[index].tolist()]
        joined = " ".join(value for value in row_values if value)
        if any(keyword in joined for keyword in ("debit", "credit")):
            return index
        if "account" in joined and any(
            keyword in joined for keyword in ("code", "name", "description")
        ):
            return index
        if "balance" in joined and "account" in joined:
            return index
    return None


def _normalize_header(value: object) -> str:
    return re.sub(r"\s+", " ", str(value).strip().lower())


def _make_unique_columns(headers: list[str], *, width: int) -> list[str]:
    unique: list[str] = []
    seen: dict[str, int] = {}
    for index in range(width):
        header = headers[index] if index < len(headers) else ""
        header = header or f"col_{index}"
        count = seen.get(header, 0)
        seen[header] = count + 1
        unique.append(header if count == 0 else f"{header}_{count}")
    return unique


def _detect_columns(columns: list[str]) -> tuple[dict[str, str], TBFormat]:
    def find(*candidates: str) -> str | None:
        for column in columns:
            for candidate in candidates:
                if candidate == column or candidate in column:
                    return column
        return None

    account_code = find("account code", "acct code", "code", "col_0")
    account_name = find("account name", "description", "name", "col_1")
    debit = find("debit", "debits", "col_2")
    credit = find("credit", "credits", "col_3")
    balance = find("balance", "net balance", "amount")

    if account_code is None or account_name is None:
        raise ParseError("Could not detect account code and account name columns.")

    if debit is not None and credit is not None:
        return {
            "account_code": account_code,
            "account_name": account_name,
            "debit": debit,
            "credit": credit,
        }, "four_column"

    if balance is not None:
        return {
            "account_code": account_code,
            "account_name": account_name,
            "balance": balance,
        }, "single_balance"

    if len(columns) >= 4:
        return {
            "account_code": columns[0],
            "account_name": columns[1],
            "debit": columns[2],
            "credit": columns[3],
        }, "four_column"

    if len(columns) == 3:
        return {
            "account_code": columns[0],
            "account_name": columns[1],
            "balance": columns[2],
        }, "single_balance"

    raise ParseError("Could not detect trial balance column layout.")


def _detect_currency(
    dataframe: pd.DataFrame,
    *,
    column_map: dict[str, str],
    functional_currency: str,
) -> tuple[str, _PerRowCurrencyResolver | None]:
    columns = list(dataframe.columns)

    currency_column = _find_currency_column(columns)
    if currency_column is not None:
        return functional_currency, _PerRowCurrencyResolver(currency_column)

    header_currency = _currency_from_headers(columns)
    if header_currency is not None:
        return header_currency, None

    symbols = _scan_currency_symbols(dataframe)
    if len(symbols) > 1:
        logger.warning(
            "Ambiguous currency symbols detected in scanned cells: %s",
            ", ".join(sorted(symbols)),
        )
        raise AmbiguousCurrencyError(frozenset(symbols))

    if len(symbols) == 1:
        return SYMBOL_TO_CURRENCY[next(iter(symbols))], None

    return functional_currency, None


def _find_currency_column(columns: list[str]) -> str | None:
    for column in columns:
        if column == "currency" or column.startswith("currency"):
            return column
    return None


def _currency_from_headers(columns: list[str]) -> str | None:
    for column in columns:
        upper = column.upper()
        for code in ISO_CURRENCIES:
            if code in upper:
                return code
    return None


def _scan_currency_symbols(dataframe: pd.DataFrame) -> set[str]:
    symbols: set[str] = set()
    scanned = 0
    for value in dataframe.to_numpy().flat:
        if scanned >= SCAN_CELL_LIMIT:
            break
        scanned += 1
        text = str(value)
        for symbol in SYMBOL_TO_CURRENCY:
            if symbol in text:
                symbols.add(symbol)
    return symbols


class _PerRowCurrencyResolver:
    def __init__(self, column_name: str) -> None:
        self._column_name = column_name

    def __call__(self, _row_index: int, series: pd.Series) -> str:
        raw = _cell_text(series, self._column_name).upper()
        if raw in ISO_CURRENCIES:
            return raw
        for symbol, code in SYMBOL_TO_CURRENCY.items():
            if symbol in raw:
                return code
        raise ParseError(f"Unsupported currency value: {raw!r}")


def _cell_text(series: pd.Series, column: str) -> str:
    value = series.get(column, "")
    return str(value).strip()


def _is_header_row(first_cell: str) -> bool:
    lowered = first_cell.lower()
    return any(keyword in lowered for keyword in HEADER_FIRST_CELL_KEYWORDS)


def _monetary_columns(column_map: dict[str, str], tb_format: TBFormat) -> tuple[str, ...]:
    if tb_format == "four_column":
        return (column_map["debit"], column_map["credit"])
    return (column_map["balance"],)


def _find_orphaned_amount(
    series: pd.Series,
    column_map: dict[str, str],
    tb_format: TBFormat,
    *,
    row_index: int,
) -> tuple[str, Decimal] | None:
    """Return the first nonzero monetary column for identifier-less rows, else None if blank."""
    for column_name in _monetary_columns(column_map, tb_format):
        raw_value = _cell_text(series, column_name)
        if raw_value == "":
            continue
        amount = parse_monetary(
            raw_value,
            row_index=row_index,
            column_name=column_name,
        )
        if amount != Decimal("0"):
            return column_name, amount
    return None


def _is_totals_row(account_name: str) -> bool:
    lowered = account_name.lower()
    return any(keyword in lowered for keyword in TOTALS_NAME_KEYWORDS)


def _validate_balanced(rows: list[TBRow]) -> None:
    total_debits = sum((row.debit for row in rows), Decimal("0"))
    total_credits = sum((row.credit for row in rows), Decimal("0"))
    if not decimal_eq(total_debits, total_credits):
        raise UnbalancedTrialBalanceError(total_debits, total_credits)
